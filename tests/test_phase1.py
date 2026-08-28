import hashlib
import hmac
import json

import pytest
from fastapi import HTTPException

from apps.api.dependencies.security import (
    validate_email_secret,
    validate_local_whatsapp_test_signature,
    validate_whatsapp_signature,
)
from apps.api.routes import integrations, test_whatsapp, whatsapp as whatsapp_admin
from services.agent_service.cx_agent import CXAgent
from services.channel_service.adapters.email_adapter import EmailAdapter
from services.channel_service.adapters.whatsapp_adapter import WhatsAppAdapter
from services.channel_service.connectors.email_sender import SMTPEmailConnector
from services.channel_service.delivery import OutboundDeliveryService
from services.crm_service.client import CRMResult
from services.intent_service.classifier import classify_intent
from services.orchestration_service.graph import OrchestrationGraph
from services.orchestration_service.router import OmnichannelRouter
from services.persistence_service.repository import SQLiteCXRepository
from services.rag_service.embeddings import SemanticEmbeddings
from services.rag_service.documents import load_knowledge_documents
from services.rag_service.rag_pipeline import RAGPipeline
from shared.schemas.intents import Intent
from shared.schemas.messages import EmailWebhookPayload, WhatsAppWebhookPayload
from shared.schemas.tickets import TicketStatus


# ── Mock LLM generators ───────────────────────────────────────────────────────

class NoLLM:
    model = "test"

    def classify_message(self, message, context):
        return None


class ValidLLM:
    """Returns a valid BFSI intent with high confidence."""
    def classify_message(self, message, context):
        return {
            "intent": "loan_status",
            "confidence": 0.9,
            "urgency": "low",
            "sentiment": "neutral",
            "reason": "Customer asked about their loan repayment status.",
        }


class UnderstatedLLM:
    """Returns an understated urgency/sentiment for a fraud message."""
    def classify_message(self, message, context):
        return {
            "intent": "fraud_report",
            "confidence": 1.0,
            "urgency": "low",
            "sentiment": "positive",
            "reason": "Customer reported fraud.",
        }


class FakeRAG:
    def answer(self, query, context):
        if "unknown" in query.lower():
            return {"answer": "manual review", "confidence": 0.0, "contexts": [], "citations": [], "llm": {}}
        context_item = {
            "text": "For loan EMI queries, customers can check their repayment schedule in the app.",
            "score": 0.91,
            "metadata": {"source": "InboxIQ_BFSI_KB.pdf:p3", "document_version": "test-v1"},
        }
        return {
            "answer": "You can check your loan repayment schedule in the app. Source: [1] InboxIQ_BFSI_KB.pdf:p3",
            "confidence": 0.91,
            "contexts": [context_item],
            "citations": [{"index": 1, "source": "InboxIQ_BFSI_KB.pdf:p3", "score": 0.91}],
            "llm": {"llm_used": False},
        }


class FakeResolutionEngine:
    """Deterministic L1/L2/L3 stub — avoids hitting real OpenSearch/Groq in tests.

    Defaults to L1 (auto-resolvable), which preserves the pre-resolution-engine
    intent-based escalation behavior exactly, since TicketCreationAgent only overrides
    on L2/L3. Pass level="L2"/"L3" to test the new override behavior explicitly.
    """

    def __init__(self, level="L1"):
        self.level = level

    def resolve_query_level(self, query, intent, sentiment):
        return {
            "intent": intent,
            "sentiment": sentiment,
            "resolution_level": self.level,
            "confidence": 0.9,
            "reason": "Test stub decision.",
        }


class FakeNeo4j:
    """Stub Neo4j client that returns canned BFSI graph data."""

    def query(self, cypher, params=None):
        if "customer_id" in cypher and "$cid" in cypher and "Loan" in cypher:
            return [{"loan_id": "L001", "loan_type": "Personal Loan", "status": "Active",
                     "amount_inr": 500000, "interest_rate": 10.5, "next_step": "Pay EMI by 5th",
                     "last_updated": "2024-01-01"}]
        if "customer_id" in cypher and "$cid" in cypher and "Claim" in cypher:
            return []
        if "phone" in cypher or "email" in cypher:
            return [{"customer_id": "CUST-001", "email": "customer@example.com",
                     "phone": "+919999999999", "city": "Mumbai"}]
        return []

    def write(self, cypher, params=None):
        pass

    def close(self):
        pass

    enabled = True


class EmptyNeo4j:
    """Active Neo4j test double with no matching customer graph data."""

    def query(self, cypher, params=None):
        return []

    def write(self, cypher, params=None):
        pass

    def close(self):
        pass

    enabled = True


class WorkbookNeo4j:
    """Neo4j test double backed by the same XLSX rows used by the graph loader."""

    enabled = True

    def __init__(self):
        import openpyxl

        workbook = openpyxl.load_workbook("data/bfsi.xlsx", data_only=True)
        self.customers = self._rows(workbook["Customer_Demographics"])
        self.loans = self._rows(workbook["Loans"])
        self.claims = self._rows(workbook["Claims"])
        self.policies = self._rows(workbook["Policies"])

    @staticmethod
    def _rows(sheet):
        headers = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
        return [dict(zip(headers, row)) for row in sheet.iter_rows(min_row=2, values_only=True)]

    def query(self, cypher, params=None):
        params = params or {}
        if "MATCH (c:Customer)" in cypher:
            identifiers = {str(params.get("id", "")), str(params.get("stripped", ""))}
            for row in self.customers:
                values = {
                    str(row.get("Mobile1") or ""),
                    str(row.get("Email1") or ""),
                    str(row.get("AlternateEmail") or ""),
                }
                if identifiers & values:
                    return [{
                        "customer_id": str(row["CRN"]),
                        "email": str(row.get("Email1") or ""),
                        "phone": str(row.get("Mobile1") or ""),
                        "city": str(row.get("City") or ""),
                        "registration_date": str(row.get("AccountOpeningDate") or ""),
                    }]
            return []

        customer_id = str(params.get("cid", ""))
        if "HAS_LOAN" in cypher:
            return [{
                "loan_id": str(row["LoanID"]),
                "loan_type": str(row["LoanType"]),
                "status": str(row["Status"]),
                "amount_inr": row["BalanceDue"],
                "interest_rate": row["InterestRate"],
                "next_step": str(row["NextStep"]),
                "last_updated": str(row["LastUpdatedDate"]),
            } for row in self.loans if str(row["CRN"]) == customer_id]
        if "HAS_CLAIM" in cypher:
            policy_type_by_id = {str(p["PolicyID"]): str(p.get("PolicyType") or "") for p in self.policies}
            return [{
                "claim_id": str(row["ClaimID"]),
                "policy_type": policy_type_by_id.get(str(row.get("PolicyID") or ""), ""),
                "claim_type": str(row["ClaimType"]),
                "status": str(row["ClaimStatus"]),
                "amount_claimed": row["AmountClaimed"],
                "amount_approved": row["AmountApproved"],
                "reason": str(row["ReasonForStatus"]),
                "last_updated": str(row["LastUpdatedDate"]),
            } for row in self.claims if str(row["CRN"]) == customer_id]
        return []

    def write(self, cypher, params=None):
        pass

    def close(self):
        pass


class Recorder:
    def __init__(self):
        self.sent = []

    def send_text(self, *args):
        self.sent.append(args)
        return {"id": "provider-message-id"}


_DEFAULT_TEST_NEO4J = object()


def graph(repository, whatsapp=None, email=None, crm=None, neo4j_client=_DEFAULT_TEST_NEO4J, resolution_engine=None):
    if neo4j_client is _DEFAULT_TEST_NEO4J:
        neo4j_client = EmptyNeo4j()
    return OrchestrationGraph(
        repository,
        agent=CXAgent(NoLLM()),
        rag=FakeRAG(),
        delivery=OutboundDeliveryService(whatsapp=whatsapp, email=email),
        crm=crm,
        neo4j_client=neo4j_client,
        resolution_engine=resolution_engine or FakeResolutionEngine(),
    )


def whatsapp_message(message_id="wamid-1", text="What is my loan EMI status?", metadata=None):
    return WhatsAppAdapter().normalize(
        WhatsAppWebhookPayload(
            from_="+919999999999",
            text=text,
            message_id=message_id,
            metadata={"provider": "whatsapp_cloud", **(metadata or {})},
        )
    )


def email_message(message_id="email-1", body="What is my loan EMI status?", metadata=None):
    return EmailAdapter().normalize(
        EmailWebhookPayload(
            from_email="customer@example.com",
            subject="Loan query",
            body=body,
            message_id=message_id,
            metadata={"provider": "email_webhook", **(metadata or {})},
        )
    )


# ── Core workflow tests ───────────────────────────────────────────────────────

def test_whatsapp_bfsi_query_resolves_with_citation_and_sends_reply():
    repo = SQLiteCXRepository(":memory:")
    sender = Recorder()
    # Non-account-specific loan question (loan_application, not loan_status) so this
    # message passes customer validation and exercises the RAG/KB fallback path.
    response = graph(repo, whatsapp=sender).run(
        whatsapp_message(text="What are the requirements for a personal loan?")
    )
    assert response.resolved is False
    assert response.workflow_status == "answer_delivered"
    assert response.citations[0]["source"] == "InboxIQ_BFSI_KB.pdf:p3"
    assert response.outbound_status == "sent"
    assert sender.sent
    ticket_step = next(entry for entry in response.workflow_trace if entry["step"] == "create_or_update_ticket")
    assert ticket_step["details"]["skipped"] is True
    evidence = repo.list_retrieval_evidence()
    assert evidence[0]["source"] == "InboxIQ_BFSI_KB.pdf:p3"


def test_unregistered_customer_blocked_from_account_specific_query():
    """A loan_status question from a sender not found in the BFSI customer graph is
    rejected before any KB/graph lookup or ticket creation, per the customer-validation gate."""
    repo = SQLiteCXRepository(":memory:")
    sender = Recorder()
    response = graph(repo, whatsapp=sender).run(whatsapp_message(text="What is my loan status?"))
    assert response.workflow_status == "customer_validation_required"
    assert response.intent == "customer_not_registered"
    assert response.ticket_id is None
    assert response.outbound_status == "sent"
    assert "registered" in response.message.lower()
    reject_step = next(entry for entry in response.workflow_trace if entry["step"] == "reject_unregistered_customer")
    assert reject_step["details"]["reason"] == "no_matching_bfsi_customer_record"


def test_registered_customer_account_query_is_not_blocked():
    """A loan_status question from a phone number that matches a real BFSI customer in
    Neo4j proceeds through the normal pipeline instead of being rejected."""
    repo = SQLiteCXRepository(":memory:")
    sender = Recorder()
    neo4j = WorkbookNeo4j()
    message = WhatsAppAdapter().normalize(WhatsAppWebhookPayload(
        from_="+917538870992",
        text="What is the status of my personal loan?",
        message_id="wamid-registered-loan",
        metadata={"provider": "test"},
    ))
    workflow = OrchestrationGraph(
        repo,
        agent=CXAgent(NoLLM()),
        rag=FakeRAG(),
        delivery=OutboundDeliveryService(whatsapp=sender),
        neo4j_client=neo4j,
        resolution_engine=FakeResolutionEngine(),
    )
    response = workflow.run(message)
    assert response.workflow_status != "customer_validation_required"
    assert response.intent == "loan_status"
    validate_step = next(entry for entry in response.workflow_trace if entry["step"] == "validate_customer")
    assert validate_step["details"]["is_registered"] is True
    assert response.outbound_status == "sent"


def test_resolution_l3_overrides_never_escalate_intent():
    """loan_status is an INFORMATIONAL_INTENT that normally never creates a ticket (Rule 3b).
    An L3 resolution-level decision must override that and force escalation anyway — this is
    the deliberate precedence documented in TicketCreationAgent._escalation_reason."""
    repo = SQLiteCXRepository(":memory:")
    sender = Recorder()
    neo4j = WorkbookNeo4j()
    message = WhatsAppAdapter().normalize(WhatsAppWebhookPayload(
        from_="+917538870992",
        text="What is the status of my personal loan?",
        message_id="wamid-l3-override",
        metadata={"provider": "test"},
    ))
    workflow = OrchestrationGraph(
        repo,
        agent=CXAgent(NoLLM()),
        rag=FakeRAG(),
        delivery=OutboundDeliveryService(whatsapp=sender),
        neo4j_client=neo4j,
        resolution_engine=FakeResolutionEngine(level="L3"),
    )
    response = workflow.run(message)
    assert response.intent == "loan_status"
    assert response.ticket_id is not None
    ticket_step = next(entry for entry in response.workflow_trace if entry["step"] == "create_or_update_ticket")
    assert "skipped" not in ticket_step["details"]
    ticket = repo.get_ticket(response.ticket_id)
    assert ticket["escalation_reason"].startswith("critical_escalation:")


def test_resolution_l1_preserves_never_escalate_intent():
    """The same loan_status query with an L1 decision falls through to the normal
    intent-based rules, which never escalate INFORMATIONAL_INTENTS — confirming L1 doesn't
    change existing behavior."""
    repo = SQLiteCXRepository(":memory:")
    sender = Recorder()
    neo4j = WorkbookNeo4j()
    message = WhatsAppAdapter().normalize(WhatsAppWebhookPayload(
        from_="+917538870992",
        text="What is the status of my personal loan?",
        message_id="wamid-l1-no-override",
        metadata={"provider": "test"},
    ))
    workflow = OrchestrationGraph(
        repo,
        agent=CXAgent(NoLLM()),
        rag=FakeRAG(),
        delivery=OutboundDeliveryService(whatsapp=sender),
        neo4j_client=neo4j,
        resolution_engine=FakeResolutionEngine(level="L1"),
    )
    response = workflow.run(message)
    assert response.intent == "loan_status"
    assert response.ticket_id is None


def test_high_risk_keyword_forces_l3_before_llm_call():
    """The deterministic safety net in services.resolution_service must force L3 for
    credible-risk language even without a real LLM/OpenSearch, and that must override a
    normally-never-escalate intent end to end."""
    repo = SQLiteCXRepository(":memory:")
    sender = Recorder()
    neo4j = WorkbookNeo4j()
    message = WhatsAppAdapter().normalize(WhatsAppWebhookPayload(
        from_="+917538870992",
        text="What is the status of my personal loan? Someone hacked my account without my permission.",
        message_id="wamid-safety-net",
        metadata={"provider": "test"},
    ))
    # No resolution_engine override here — exercises the REAL services.resolution_service
    # safety-net keyword check (no injected fake), confirming it works end to end.
    workflow = OrchestrationGraph(
        repo,
        agent=CXAgent(NoLLM()),
        rag=FakeRAG(),
        delivery=OutboundDeliveryService(whatsapp=sender),
        neo4j_client=neo4j,
    )
    response = workflow.run(message)
    assert response.ticket_id is not None
    ticket = repo.get_ticket(response.ticket_id)
    assert ticket["escalation_reason"].startswith("critical_escalation:")


def test_distinct_l3_intents_create_distinct_team_tickets():
    repo = SQLiteCXRepository(":memory:")
    sender = Recorder()
    workflow = OrchestrationGraph(
        repo,
        agent=CXAgent(NoLLM()),
        rag=FakeRAG(),
        delivery=OutboundDeliveryService(whatsapp=sender),
        neo4j_client=WorkbookNeo4j(),
        resolution_engine=FakeResolutionEngine(level="L3"),
    )

    fraud = workflow.run(
        WhatsAppAdapter().normalize(WhatsAppWebhookPayload(
            from_="+917538870992",
            text="Someone hacked my account and stole money. This is fraud, help immediately.",
            message_id="l3-fraud-separate-ticket",
            metadata={"provider": "whatsapp_cloud"},
        ))
    )
    default = workflow.run(
        WhatsAppAdapter().normalize(WhatsAppWebhookPayload(
            from_="+917538870992",
            text="I received a loan default notice but I already paid my EMI.",
            message_id="l3-default-separate-ticket",
            metadata={"provider": "whatsapp_cloud"},
        ))
    )

    fraud_ticket = repo.get_ticket(fraud.ticket_id)
    default_ticket = repo.get_ticket(default.ticket_id)

    assert fraud_ticket["ticket_id"] != default_ticket["ticket_id"]
    assert fraud_ticket["intent"] == "fraud_report"
    assert fraud_ticket["assigned_team"] == "fraud_and_disputes"
    assert default_ticket["intent"] == "loan_default_notice"
    assert default_ticket["assigned_team"] == "collections"


def test_distinct_l3_fraud_incidents_create_distinct_tickets():
    repo = SQLiteCXRepository(":memory:")
    sender = Recorder()
    workflow = OrchestrationGraph(
        repo,
        agent=CXAgent(NoLLM()),
        rag=FakeRAG(),
        delivery=OutboundDeliveryService(whatsapp=sender),
        neo4j_client=WorkbookNeo4j(),
        resolution_engine=FakeResolutionEngine(level="L3"),
    )

    takeover = workflow.run(
        WhatsAppAdapter().normalize(WhatsAppWebhookPayload(
            from_="+917538870992",
            text="Someone hacked my account and transferred money without my permission.",
            message_id="l3-fraud-account-takeover",
            metadata={"provider": "whatsapp_cloud"},
        ))
    )
    phishing = workflow.run(
        WhatsAppAdapter().normalize(WhatsAppWebhookPayload(
            from_="+917538870992",
            text="I got a phishing link, entered my banking details, and now I cannot access my account",
            message_id="l3-fraud-phishing-compromise",
            metadata={"provider": "whatsapp_cloud"},
        ))
    )
    repeated_takeover = workflow.run(
        WhatsAppAdapter().normalize(WhatsAppWebhookPayload(
            from_="+917538870992",
            text="Someone hacked my account and transferred money without my permission.",
            message_id="l3-fraud-account-takeover-repeat",
            metadata={"provider": "whatsapp_cloud"},
        ))
    )

    takeover_ticket = repo.get_ticket(takeover.ticket_id)
    phishing_ticket = repo.get_ticket(phishing.ticket_id)

    assert takeover.ticket_id != phishing.ticket_id
    assert repeated_takeover.ticket_id == takeover.ticket_id
    assert takeover_ticket["intent"] == "fraud_report"
    assert phishing_ticket["intent"] == "fraud_report"
    assert takeover_ticket["assigned_team"] == "fraud_and_disputes"
    assert phishing_ticket["assigned_team"] == "fraud_and_disputes"
    assert takeover_ticket["metadata"]["ticket_scope"] == "fraud_report:account_takeover_funds_stolen"
    assert phishing_ticket["metadata"]["ticket_scope"] == "fraud_report:phishing_credential_compromise"


def test_email_complaint_escalates_and_sends_reply():
    repo = SQLiteCXRepository(":memory:")
    sender = Recorder()
    response = graph(repo, email=sender).run(
        email_message(body="This is an unacceptable complaint. I need a human agent immediately.")
    )
    assert response.resolved is False
    assert response.intent in {"complaint", "human_escalation"}
    assert response.ticket_id
    assert response.outbound_status == "sent"
    assert sender.sent
    ticket_step = next(entry for entry in response.workflow_trace if entry["step"] == "create_or_update_ticket")
    assert "skipped" not in ticket_step["details"]


def test_outbound_failure_error_is_returned_in_channel_response():
    class FailingSender:
        def send_text(self, *args):
            raise RuntimeError("provider rejected recipient")

    repo = SQLiteCXRepository(":memory:")
    response = graph(repo, whatsapp=FailingSender()).run(whatsapp_message())

    assert response.outbound_status == "failed"
    assert response.outbound_error == "provider rejected recipient"


def test_duplicate_message_returns_original_response_without_second_send():
    repo = SQLiteCXRepository(":memory:")
    sender = Recorder()
    workflow = graph(repo, whatsapp=sender)
    first = workflow.run(whatsapp_message())
    second = workflow.run(whatsapp_message())
    assert second.duplicate is True
    assert first.conversation_id == second.conversation_id
    assert len(sender.sent) == 1


def test_customer_identity_resolution_links_email_and_whatsapp():
    repo = SQLiteCXRepository(":memory:")
    workflow = graph(repo)
    first = workflow.run(whatsapp_message(metadata={"linked_email": "customer@example.com"}))
    second = workflow.run(email_message())
    assert first.customer_id == second.customer_id
    assert first.conversation_id == second.conversation_id


def test_multi_turn_context_is_persisted():
    repo = SQLiteCXRepository(":memory:")
    workflow = graph(repo)
    response = workflow.run(whatsapp_message())
    workflow.run(whatsapp_message(message_id="wamid-2", text="Can you check my EMI again?"))
    conversation = repo.get_conversation(response.conversation_id)
    assert len(conversation["turns"]) == 4
    assert "EMI" in conversation["summary"]


def test_customer_message_can_resolve_active_ticket_without_rag():
    repo = SQLiteCXRepository(":memory:")
    sender = Recorder()
    workflow = graph(repo, whatsapp=sender)
    opened = workflow.run(
        whatsapp_message(message_id="wamid-open", text="This is a complaint. Please connect me to a human.")
    )
    closed = workflow.run(
        whatsapp_message(message_id="wamid-close", text="close the ticket as query is resolved, thanks")
    )

    assert opened.ticket_id
    assert closed.ticket_id == opened.ticket_id
    assert closed.intent == "ticket_resolution"
    assert closed.resolved is True
    assert closed.workflow_status == "ticket_closed"
    assert closed.retrieval_backend == "not_required"
    assert closed.rag_contexts == []
    assert repo.get_ticket(opened.ticket_id)["status"] == TicketStatus.RESOLVED.value
    assert "marked as resolved" in closed.message
    # The customer has an open case, so check_has_open_case routes into the ticket
    # branch; detect_ticket_action then finds a close request and select_ticket_to_resolve
    # picks which one (only one candidate here, so no clarification is needed).
    assert [entry["step"] for entry in closed.workflow_trace] == [
        "receive_message",
        "resolve_identity",
        "load_conversation_context",
        "check_has_open_case",
        "detect_ticket_action",
        "select_ticket_to_resolve",
        "resolve_ticket",
        "send_outbound_reply",
        "persist_audit_events",
    ]


def test_has_open_case_gate_routes_on_customer_state_not_message_content():
    """The gate answers "does this customer have a case", not "is this a close request".

    A customer with an open ticket asking something unrelated must still take the ticket
    branch (has_open_case=1) — the earlier version of this node keyed on the message being
    a resolution, so a customer with three open cases asking a fresh question read 0.
    """
    repo = SQLiteCXRepository(":memory:")
    workflow = graph(repo)

    # Distinct message ids: whatsapp_message() defaults to "wamid-1", and a repeat of an
    # already-seen provider message id is correctly suppressed as a duplicate delivery.
    first = workflow.run(
        whatsapp_message(message_id="case-gate-1", text="Someone used my card without my permission.")
    )
    steps = [entry["step"] for entry in first.workflow_trace]
    gate = [e for e in first.workflow_trace if e["step"] == "check_has_open_case"][0]
    # No case existed when this turn arrived, so the ticket branch is skipped outright.
    assert gate["details"]["has_open_case"] == 0
    assert "detect_ticket_action" not in steps
    assert first.ticket_id

    second = workflow.run(
        whatsapp_message(message_id="case-gate-2", text="What are the requirements for a personal loan?")
    )
    gate2 = [e for e in second.workflow_trace if e["step"] == "check_has_open_case"][0]
    steps2 = [entry["step"] for entry in second.workflow_trace]
    # Same customer now HAS an open case, so the gate is 1 even though this message is an
    # ordinary question — and detect_ticket_action correctly declines to close anything.
    assert gate2["details"]["has_open_case"] == 1
    assert "detect_ticket_action" in steps2
    assert "resolve_ticket" not in steps2
    assert second.resolved is False


def test_select_ticket_asks_which_when_two_open_tickets_share_a_scope():
    """Two open cases of the same kind must produce a question, not a guess."""
    from services.agent_service.orchestration_agents import TicketCreationAgent
    from services.ticket_service.ticket_manager import TicketManager

    repo = SQLiteCXRepository(":memory:")
    agent = TicketCreationAgent(TicketManager(repo))

    def ticket(tid, intent, scope):
        return {"ticket_id": tid, "intent": intent, "metadata": {"ticket_scope": scope}}

    card_a = ticket("tkt_aaaaaaaaaaaa", "transaction_dispute", "transaction_dispute:card")
    card_b = ticket("tkt_bbbbbbbbbbbb", "transaction_dispute", "transaction_dispute:card")
    upi = ticket("tkt_cccccccccccc", "transaction_dispute", "transaction_dispute:upi")

    def select(text, active, owned):
        return agent.select_ticket(
            whatsapp_message(text=text), {"active_ticket": active, "customer_tickets": owned}
        )

    # One candidate of that kind -> resolve it outright.
    assert select("yes resolved", card_a, [card_a]).target_ticket_id == "tkt_aaaaaaaaaaaa"

    # Two of the SAME scope and no id named -> ask, and resolve nothing.
    ambiguous = select("yes resolved", card_a, [card_a, card_b])
    assert ambiguous.needs_clarification is True
    assert ambiguous.target_ticket_id is None
    assert {c["ticket_id"] for c in ambiguous.candidates} == {"tkt_aaaaaaaaaaaa", "tkt_bbbbbbbbbbbb"}

    # A card dispute and a UPI dispute are different matters despite one intent.
    assert select("yes resolved", card_a, [card_a, upi]).target_ticket_id == "tkt_aaaaaaaaaaaa"

    # An explicitly named id wins over the ambiguity.
    named = select("please close tkt_bbbbbbbbbbbb", card_a, [card_a, card_b])
    assert named.target_ticket_id == "tkt_bbbbbbbbbbbb"

    # An id the customer does NOT own is ignored rather than honoured.
    foreign = select("close tkt_ffffffffffff", card_a, [card_a, card_b])
    assert foreign.needs_clarification is True
    assert foreign.target_ticket_id is None


def test_restart_persistence(tmp_path):
    database = str(tmp_path / "phase1.db")
    first_repo = SQLiteCXRepository(database)
    response = graph(first_repo).run(whatsapp_message())
    second_repo = SQLiteCXRepository(database)
    conversation = second_repo.get_conversation(response.conversation_id)
    assert conversation
    assert len(conversation["turns"]) == 2
    assert second_repo.list_audit_events(response.correlation_id)


def test_whatsapp_cloud_webhook_e2e_preserves_channel_and_sends_reply(monkeypatch):
    from fastapi.testclient import TestClient

    from apps.api.main import app
    from apps.api.routes import webhooks

    repo = SQLiteCXRepository(":memory:")
    sender = Recorder()
    router = OmnichannelRouter(graph(repo, whatsapp=sender))
    monkeypatch.setattr(webhooks, "get_router", lambda: router)
    monkeypatch.setenv("WHATSAPP_APP_SECRET", "e2e-whatsapp-secret")

    payload = {
        "entry": [
            {
                "changes": [
                    {
                        "value": {
                            "contacts": [
                                {"wa_id": "918555870077", "profile": {"name": "Phase 1 Tester"}}
                            ],
                            "messages": [
                                {
                                    "from": "918555870077",
                                    "id": "wamid-e2e-kyc",
                                    "timestamp": "1710000000",
                                    "type": "text",
                                    "text": {"body": "How can I update my KYC?"},
                                }
                            ],
                        }
                    }
                ]
            }
        ]
    }
    body = json.dumps(payload, separators=(",", ":")).encode()
    signature = "sha256=" + hmac.new(b"e2e-whatsapp-secret", body, hashlib.sha256).hexdigest()

    response = TestClient(app).post(
        "/integrations/whatsapp/webhook",
        content=body,
        headers={"x-hub-signature-256": signature, "content-type": "application/json"},
    )

    assert response.status_code == 202
    result = response.json()["messages"][0]
    assert response.json()["messages_received"] == 1
    assert result["outbound_status"] == "sent"
    assert result["workflow_trace"][0]["step"] == "receive_message"
    assert sender.sent[0][0] == "918555870077"

    conversation = repo.get_conversation(result["conversation_id"])
    assert [turn["channel"] for turn in conversation["turns"]] == ["whatsapp", "whatsapp"]
    assert conversation["turns"][0]["external_message_id"] == "wamid-e2e-kyc"
    assert conversation["turns"][1]["delivery_status"] == "sent"


def test_email_webhook_e2e_preserves_channel_creates_ticket_and_sends_reply(monkeypatch):
    from fastapi.testclient import TestClient

    from apps.api.main import app
    from apps.api.routes import webhooks

    repo = SQLiteCXRepository(":memory:")
    sender = Recorder()
    router = OmnichannelRouter(graph(repo, email=sender))
    monkeypatch.setattr(webhooks, "get_router", lambda: router)
    monkeypatch.setenv("EMAIL_WEBHOOK_SECRET", "e2e-email-secret")

    response = TestClient(app).post(
        "/integrations/email/webhook",
        json={
            "from_email": "customer@example.com",
            "subject": "Loan support",
            "body": "What is the status of my loan LN001? Please connect me to a human agent.",
            "message_id": "email-e2e-human-agent",
            "metadata": {"source": "phase1-e2e"},
        },
        headers={"x-email-webhook-secret": "e2e-email-secret"},
    )

    assert response.status_code == 200
    result = response.json()
    assert result["outbound_status"] == "sent"
    assert result["ticket_id"]
    assert result["workflow_status"] == "human_follow_up"
    assert sender.sent[0][0] == "customer@example.com"

    conversation = repo.get_conversation(result["conversation_id"])
    assert [turn["channel"] for turn in conversation["turns"]] == ["email", "email"]
    assert conversation["turns"][0]["external_message_id"] == "email-e2e-human-agent"
    assert conversation["turns"][1]["ticket_id"] == result["ticket_id"]


# ── BFSI intent classification tests ─────────────────────────────────────────

def test_intent_classifier_supports_bfsi_intents():
    assert classify_intent("Please report a fraud transaction on my account").intent == Intent.FRAUD_REPORT
    assert classify_intent("What is my outstanding loan EMI?").intent == Intent.LOAN_STATUS
    assert classify_intent("I want to speak to a human representative").intent == Intent.HUMAN_ESCALATION
    assert classify_intent("Block my lost credit card immediately").intent == Intent.CARD_MANAGEMENT
    assert classify_intent("I need to file an insurance claim for my accident").intent == Intent.INSURANCE_CLAIM
    assert classify_intent("What is the status of my existing claim?").intent == Intent.CLAIM_STATUS
    assert classify_intent("How do I apply for a home loan?").intent == Intent.LOAN_APPLICATION
    assert classify_intent("I received a loan default notice but I already paid my EMI").intent == Intent.LOAN_DEFAULT_NOTICE
    assert classify_intent("What is SIP?").intent == Intent.GENERAL_INQUIRY
    assert classify_intent("What is an ELSS scheme?").intent == Intent.GENERAL_INQUIRY


def test_llm_intent_result_accepts_model_reason():
    result = CXAgent(ValidLLM()).analyze("What is the status of my loan repayment?")
    assert result.intent == Intent.LOAN_STATUS
    assert result.reason == "Customer asked about their loan repayment status."
    assert result.analysis_source == "groq_llm"


def test_groq_generator_records_local_llm_usage(monkeypatch, tmp_path):
    from services.observability_service import llm_observation_context
    from services.rag_service.groq_generator import GroqGenerator

    class Usage:
        prompt_tokens = 120
        completion_tokens = 30
        total_tokens = 150

    class Message:
        content = "Observed answer"

    class Choice:
        message = Message()

    class Response:
        choices = [Choice()]
        usage = Usage()

    class FakeCompletions:
        def create(self, **kwargs):
            return Response()

    class FakeChat:
        completions = FakeCompletions()

    class FakeGroq:
        chat = FakeChat()

    db_path = tmp_path / "llm_usage.db"
    monkeypatch.setenv("DATABASE_PATH", str(db_path))
    monkeypatch.setenv("LLM_OBSERVABILITY_ENABLED", "true")
    monkeypatch.setenv("LANGFUSE_ENABLED", "false")
    # Pin the model rather than inheriting GROQ_MODEL from the ambient .env: the rate
    # table below is keyed by model name, so a generator built on a different model
    # costs out at 0 and the estimated_cost_usd assertion fails for the wrong reason.
    monkeypatch.setenv("GROQ_MODEL", "llama-3.1-8b-instant")
    monkeypatch.setenv("LLM_COST_RATES_JSON", '{"llama-3.1-8b-instant":{"input":0.05,"output":0.08}}')

    generator = GroqGenerator()
    generator.api_key = "test-key"
    generator._client = FakeGroq()

    with llm_observation_context(
        correlation_id="corr_llm_test",
        conversation_id="conv_llm_test",
        customer_id="cust_llm_test",
        message_id="msg_llm_test",
        channel="whatsapp",
        agent="query_resolution_agent",
        intent=Intent.GENERAL_INQUIRY.value,
    ):
        result = generator._generate(
            system_prompt="system",
            user_prompt="customer question",
            operation="answer_generation",
            metadata={"context_count": 2},
        )

    repo = SQLiteCXRepository(str(db_path))
    events = repo.list_llm_usage_events()
    summary = repo.get_llm_usage_summary(days=7)

    assert result["text"] == "Observed answer"
    assert len(events) == 1
    assert events[0]["correlation_id"] == "corr_llm_test"
    assert events[0]["operation"] == "answer_generation"
    assert events[0]["agent"] == "query_resolution_agent"
    assert events[0]["intent"] == Intent.GENERAL_INQUIRY.value
    assert events[0]["prompt_tokens"] == 120
    assert events[0]["completion_tokens"] == 30
    assert events[0]["total_tokens"] == 150
    assert events[0]["estimated_cost_usd"] > 0
    assert events[0]["metadata"]["context_count"] == 2
    assert summary["totals"]["calls"] == 1
    assert summary["by_operation"][0]["operation"] == "answer_generation"


def test_llm_intent_guardrails_raise_understated_sentiment_and_urgency():
    result = CXAgent(UnderstatedLLM()).analyze(
        "Someone hacked my account and stole all my money. This is fraud! Help immediately!"
    )
    assert result.intent == Intent.FRAUD_REPORT
    assert result.sentiment == "negative"
    assert result.urgency.value == "high"
    assert "Deterministic guardrails raised: sentiment, urgency." in result.reason


# ── Neo4j enrichment tests ────────────────────────────────────────────────────

def test_neo4j_context_enriches_intent_classification():
    repo = SQLiteCXRepository(":memory:")
    fake_neo4j = FakeNeo4j()
    response = graph(repo, neo4j_client=fake_neo4j).run(
        whatsapp_message(text="What is my loan status?")
    )
    assert response.intent in {
        "loan_status", "general_inquiry", "loan_application", "loan_default_notice"
    }


def test_query_resolution_agent_routes_transactional_intent_to_neo4j():
    from services.agent_service.orchestration_agents import QueryResolutionAgent
    from services.channel_service.adapters.whatsapp_adapter import WhatsAppAdapter
    from shared.schemas.messages import WhatsAppWebhookPayload

    agent = QueryResolutionAgent(neo4j_client=FakeNeo4j(), resolution_engine=FakeResolutionEngine())
    msg = WhatsAppAdapter().normalize(
        WhatsAppWebhookPayload(from_="+919999999999", text="What is my loan status?",
                               message_id="test-1", metadata={"provider": "test"})
    )
    context = {"graph_context": {"customer_id": "CUST-001"}}
    resolution = agent.run(msg, context, intent="loan_status")
    assert resolution.retrieval_backend == "neo4j_graph"
    assert "loan" in resolution.answer.lower() or "L001" in resolution.answer


def test_query_resolution_agent_routes_general_inquiry_to_rag():
    from services.agent_service.orchestration_agents import QueryResolutionAgent
    from services.channel_service.adapters.whatsapp_adapter import WhatsAppAdapter
    from shared.schemas.messages import WhatsAppWebhookPayload

    fake_rag = FakeRAG()
    agent = QueryResolutionAgent(rag=fake_rag, neo4j_client=FakeNeo4j(), resolution_engine=FakeResolutionEngine())
    msg = WhatsAppAdapter().normalize(
        WhatsAppWebhookPayload(from_="+919999999999", text="What documents do I need for KYC?",
                               message_id="test-2", metadata={"provider": "test"})
    )
    resolution = agent.run(msg, {}, intent="kyc_update")
    # kyc_update is not a TRANSACTIONAL_INTENT → should use RAG
    assert resolution.retrieval_backend != "neo4j_graph"


def test_general_inquiry_resolution_memory_does_not_override_kb_rag():
    from services.agent_service.orchestration_agents import QueryResolutionAgent

    class MemoryNeo4j(EmptyNeo4j):
        def query(self, cypher, params=None):
            if "ResolutionMemory" in cypher:
                return [{
                    "resolution": "Shared the latest account features and applicable charges for your reference.",
                    "times_reused": 7,
                    "verified": True,
                    "query_pattern": "What are the current features and charges applicable on my account?",
                }]
            return super().query(cypher, params)

    agent = QueryResolutionAgent(
        rag=FakeRAG(),
        neo4j_client=MemoryNeo4j(),
        resolution_engine=FakeResolutionEngine(),
    )
    resolution = agent.run(
        whatsapp_message(text="What is SIP?"),
        {},
        intent=Intent.GENERAL_INQUIRY.value,
    )

    assert resolution.retrieval_backend != "resolution_memory_cache"
    assert "account features and applicable charges" not in resolution.answer.lower()


def test_process_intents_never_route_to_customer_graph():
    from services.agent_service.orchestration_agents import QueryResolutionAgent

    fake_rag = FakeRAG()
    agent = QueryResolutionAgent(rag=fake_rag, neo4j_client=FakeNeo4j(), resolution_engine=FakeResolutionEngine())
    context = {"graph_context": {"customer_id": "CUST-001"}}

    claim = agent.run(
        whatsapp_message(text="How do I file a health insurance claim?"),
        context,
        intent=Intent.INSURANCE_CLAIM.value,
    )
    loan = agent.run(
        whatsapp_message(text="How do I apply for a home loan?"),
        context,
        intent=Intent.LOAN_APPLICATION.value,
    )

    assert claim.retrieval_backend != "neo4j_graph"
    assert loan.retrieval_backend != "neo4j_graph"


def test_five_question_kb_and_graph_e2e_matrix():
    class EmptyStore:
        def similarity_search(self, query, k):
            return []

    class NoGeneration:
        model = "test"

        def generate_answer(self, query, contexts, conversation_context):
            return {"text": "", "model": self.model, "llm_used": False}

    rag = RAGPipeline(store=EmptyStore(), generator=NoGeneration())
    neo4j = WorkbookNeo4j()

    cases = [
        {
            "message": WhatsAppAdapter().normalize(WhatsAppWebhookPayload(
                from_="917700920746",
                text="How do I file a health insurance claim?",
                message_id="matrix-kb-health-claim",
                metadata={"provider": "test"},
            )),
            "intent": Intent.INSURANCE_CLAIM.value,
            "backend": "keyword_fallback",
            "source_page": ":p2",
            "expected": ("cashless", "reimbursement"),
            "forbidden": "No insurance claims",
        },
        {
            "message": EmailAdapter().normalize(EmailWebhookPayload(
                from_email="digvijayyadav48@gmail.com",
                subject="Home loan application",
                body="How do I apply for a home loan?",
                message_id="matrix-kb-home-loan",
                metadata={"provider": "test"},
            )),
            "intent": Intent.LOAN_APPLICATION.value,
            "backend": "keyword_fallback",
            "source_page": ":p1",
            "expected": ("application form", "property documents"),
            "forbidden": "No active loan",
        },
        {
            "message": WhatsAppAdapter().normalize(WhatsAppWebhookPayload(
                from_="917700920746",
                text="How do I report a lost or stolen credit card?",
                message_id="matrix-kb-lost-card",
                metadata={"provider": "test"},
            )),
            "intent": Intent.CARD_MANAGEMENT.value,
            "backend": "keyword_fallback",
            "source_page": ":p1",
            "expected": ("block", "replacement"),
            "forbidden": "No active",
        },
        {
            "message": WhatsAppAdapter().normalize(WhatsAppWebhookPayload(
                from_="919876510100",
                text="What is the status of my home loan application?",
                message_id="matrix-graph-loan-status",
                metadata={"provider": "test"},
            )),
            "intent": Intent.LOAN_STATUS.value,
            "backend": "neo4j_graph",
            "source_page": None,
            "expected": ("LN001", "LN016", "Under Review"),
            "forbidden": "application form",
        },
        {
            "message": EmailAdapter().normalize(EmailWebhookPayload(
                from_email="fathima.devasahayam@ganitinc.com",
                subject="Claim status",
                body="What is the status of my existing insurance claim?",
                message_id="matrix-graph-claim-status",
                metadata={"provider": "test"},
            )),
            "intent": Intent.CLAIM_STATUS.value,
            "backend": "neo4j_graph",
            "source_page": None,
            "expected": ("CLM001", "CLM016", "Under Review"),
            "forbidden": "cashless claims",
        },
    ]

    for case in cases:
        repository = SQLiteCXRepository(":memory:")
        sender = Recorder()
        delivery = (
            OutboundDeliveryService(whatsapp=sender)
            if case["message"].channel.value == "whatsapp"
            else OutboundDeliveryService(email=sender)
        )
        workflow = OrchestrationGraph(
            repository,
            agent=CXAgent(NoLLM()),
            rag=rag,
            delivery=delivery,
            neo4j_client=neo4j,
            resolution_engine=FakeResolutionEngine(),
        )
        response = workflow.run(case["message"])

        assert response.intent == case["intent"]
        assert response.retrieval_backend == case["backend"]
        assert all(text.lower() in response.message.lower() for text in case["expected"])
        assert case["forbidden"].lower() not in response.message.lower()
        assert response.outbound_status == "sent"
        assert sender.sent
        if case["source_page"]:
            assert response.citations
            assert case["source_page"] in response.citations[0]["source"]
        else:
            assert response.citations[0]["source"] == "neo4j_customer_graph"


# ── RAG / knowledge base tests ────────────────────────────────────────────────

def test_low_confidence_retrieval_creates_ticket():
    repo = SQLiteCXRepository(":memory:")
    response = graph(repo).run(whatsapp_message(text="unknown question xyz"))
    assert response.ticket_id
    assert repo.get_ticket(response.ticket_id)


def test_investment_faqs_are_l1_kb_answers_without_tickets():
    class EmptyStore:
        def similarity_search(self, query, k):
            return []

    class NoGeneration:
        model = "test"

        def generate_answer(self, query, contexts, conversation_context):
            return {"text": "", "model": self.model, "llm_used": False}

    repo = SQLiteCXRepository(":memory:")
    workflow = OrchestrationGraph(
        repo,
        agent=CXAgent(NoLLM()),
        rag=RAGPipeline(store=EmptyStore(), generator=NoGeneration()),
        delivery=OutboundDeliveryService(whatsapp=Recorder()),
        neo4j_client=EmptyNeo4j(),
        resolution_engine=FakeResolutionEngine(),
    )

    sip = workflow.run(whatsapp_message(message_id="sip-l1", text="What is SIP?"))
    elss = workflow.run(whatsapp_message(message_id="elss-l1", text="What is an ELSS scheme?"))

    assert sip.ticket_id is None
    assert "systematic investment plan" in sip.message.lower()
    assert sip.retrieval_backend == "keyword_fallback"
    assert elss.ticket_id is None
    assert "equity linked savings scheme" in elss.message.lower()
    assert elss.retrieval_backend == "keyword_fallback"


def test_customer_answer_kb_documents_are_knowledge_base_type():
    """PDF KB docs should all have doc_type=knowledge_base."""
    documents = load_knowledge_documents()
    assert documents
    assert {document.metadata["doc_type"] for document in documents} == {"knowledge_base"}
    assert all(document.metadata["source"].split(":p", 1)[0].endswith(".pdf") for document in documents)


def test_rag_discards_non_kb_contexts_before_generation():
    class UnsafeStore:
        def similarity_search(self, query, k):
            return [
                {
                    "text": "Customer: Example Name | Loan: L123",
                    "score": 0.99,
                    "metadata": {"source": "ticket-export.json", "doc_type": "ticket_history"},
                },
                {
                    "text": "Insurance claim process requires form IC-01.",
                    "score": 0.8,
                    "metadata": {"source": "InboxIQ_BFSI_KB.pdf:p5", "doc_type": "knowledge_base"},
                },
            ]

    class ContextRecorder:
        def generate_answer(self, query, contexts, conversation_context):
            assert [c["metadata"]["source"] for c in contexts] == ["InboxIQ_BFSI_KB.pdf:p5"]
            return {"text": "Approved answer.", "llm_used": True}

    answer = RAGPipeline(store=UnsafeStore(), generator=ContextRecorder()).answer("insurance claim process")
    assert [c["metadata"]["source"] for c in answer["contexts"]] == ["InboxIQ_BFSI_KB.pdf:p5"]


def test_rag_uses_pdf_keyword_fallback_when_vector_store_is_empty():
    class EmptyStore:
        def similarity_search(self, query, k):
            return []

    class NoGeneration:
        def generate_answer(self, query, contexts, conversation_context):
            return {"text": "", "model": "test", "llm_used": False}

    answer = RAGPipeline(store=EmptyStore(), generator=NoGeneration()).answer("credit card block")
    assert answer["contexts"]
    assert answer["retrieval_backend"] == "keyword_fallback"
    assert answer["citations"][0]["source"].startswith("InboxIQ_BFSI_KB.pdf")


# ── Security / auth tests ─────────────────────────────────────────────────────

def test_rag_reranks_stolen_card_pdf_faq_above_weak_vector_hit():
    class WeakVectorStore:
        def similarity_search(self, query, k):
            return [{
                "text": "BFSI Knowledge Base FAQs Banking Services Q: How can I open a new savings account?",
                "score": 0.5,
                "metadata": {"source": "InboxIQ_BFSI_KB.pdf:p1", "doc_type": "knowledge_base"},
            }]

    class ContextRecorder:
        def generate_answer(self, query, contexts, conversation_context):
            assert "lost or stolen" in contexts[0]["text"].lower()
            assert "block your card" in contexts[0]["text"].lower()
            return {
                "text": "Block the card immediately through the hotline, mobile app, or internet banking. [1]",
                "model": "test",
                "llm_used": True,
            }

    answer = RAGPipeline(store=WeakVectorStore(), generator=ContextRecorder()).answer(
        "what should I do, my card is stolen"
    )
    assert answer["retrieval_backend"] == "hybrid_keyword_rerank"
    assert "lost or stolen" in answer["contexts"][0]["text"].lower()


def test_email_authentication_failure(monkeypatch):
    monkeypatch.setenv("EMAIL_WEBHOOK_SECRET", "expected")
    with pytest.raises(HTTPException) as exc:
        validate_email_secret("wrong")
    assert exc.value.status_code == 401


def test_gmail_smtp_status_and_send(monkeypatch):
    sent = {}

    class FakeSMTP:
        def __init__(self, host, port, timeout):
            sent["host"] = host
            sent["port"] = port
            sent["timeout"] = timeout

        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc, traceback):
            return None

        def ehlo(self):
            sent["ehlo"] = sent.get("ehlo", 0) + 1

        def starttls(self):
            sent["starttls"] = True

        def login(self, username, password):
            sent["username"] = username
            sent["password"] = password

        def send_message(self, message):
            sent["to"] = message["To"]
            sent["subject"] = message["Subject"]
            return {}

    monkeypatch.setenv("SMTP_HOST", "smtp.gmail.com")
    monkeypatch.setenv("SMTP_PORT", "587")
    monkeypatch.setenv("SMTP_FROM_EMAIL", "support@example.com")
    monkeypatch.setenv("SMTP_USERNAME", "support@example.com")
    monkeypatch.setenv("SMTP_PASSWORD", "app-password")
    monkeypatch.setenv("SMTP_USE_TLS", "true")
    monkeypatch.setenv("SMTP_USE_SSL", "false")
    monkeypatch.setattr("smtplib.SMTP", FakeSMTP)

    connector = SMTPEmailConnector()
    assert connector.status()["gmail_ready"] is True
    result = connector.send_text("customer@example.com", "Test", "Hello")

    assert result["status"] == "sent"
    assert sent["host"] == "smtp.gmail.com"
    assert sent["port"] == 587
    assert sent["starttls"] is True
    assert sent["username"] == "support@example.com"
    assert sent["to"] == "customer@example.com"


def test_whatsapp_signature_validation(monkeypatch):
    monkeypatch.setenv("WHATSAPP_APP_SECRET", "secret")
    body = b'{"entry":[]}'
    signature = "sha256=" + hmac.new(b"secret", body, hashlib.sha256).hexdigest()
    validate_whatsapp_signature(body, signature)
    with pytest.raises(HTTPException):
        validate_whatsapp_signature(body, "sha256=wrong")


def test_local_whatsapp_test_signature_is_gated(monkeypatch):
    monkeypatch.setenv("WHATSAPP_LOCAL_TEST_MODE", "false")
    monkeypatch.setenv("WHATSAPP_TEST_SIGNATURE", "local-secret")
    with pytest.raises(HTTPException) as exc:
        validate_local_whatsapp_test_signature("local-secret")
    assert exc.value.status_code == 404

    monkeypatch.setenv("WHATSAPP_LOCAL_TEST_MODE", "true")
    validate_local_whatsapp_test_signature("local-secret")
    with pytest.raises(HTTPException) as exc:
        validate_local_whatsapp_test_signature("wrong")
    assert exc.value.status_code == 401


def test_local_whatsapp_inbound_uses_mock_outbound_provider(monkeypatch):
    monkeypatch.setenv("WHATSAPP_LOCAL_TEST_MODE", "true")
    repo = SQLiteCXRepository(":memory:")
    workflow = graph(repo)
    message = whatsapp_message(metadata={"provider": "whatsapp_local_test"})
    response = workflow.run(message)
    events = repo.list_audit_events(response.correlation_id)
    outbound = next(event for event in events if event["event_type"] == "outbound_sent")
    assert outbound["details"]["provider_response"]["provider"] == "whatsapp_local_test"


def test_local_whatsapp_inbound_can_request_meta_outbound_provider(monkeypatch):
    captured = {}

    def fake_handle_whatsapp_message(payload):
        captured["provider"] = payload.metadata["provider"]
        captured["outbound_provider"] = payload.metadata["outbound_provider"]

        class FakeResponse:
            def model_dump(self, mode=None):
                return {"ok": True}

        return FakeResponse()

    monkeypatch.setenv("WHATSAPP_LOCAL_TEST_MODE", "true")
    monkeypatch.setenv("WHATSAPP_TEST_SIGNATURE", "local-secret")
    monkeypatch.setenv("WHATSAPP_ACCESS_TOKEN", "meta-token")
    monkeypatch.setenv("WHATSAPP_PHONE_NUMBER_ID", "phone-id")
    monkeypatch.setattr(test_whatsapp, "handle_whatsapp_message", fake_handle_whatsapp_message)

    result = test_whatsapp.simulate_whatsapp_inbound(
        test_whatsapp.LocalWhatsAppInbound(
            from_="919999999999",
            text="Where is my order?",
            outbound_provider="meta",
        ),
        x_test_whatsapp_signature="local-secret",
    )

    assert result == {"ok": True}
    assert captured == {"provider": "whatsapp_cloud", "outbound_provider": "meta"}


def test_local_whatsapp_manual_send_writes_audit_event(monkeypatch):
    monkeypatch.setenv("WHATSAPP_LOCAL_TEST_MODE", "true")
    monkeypatch.setenv("WHATSAPP_TEST_SIGNATURE", "local-secret")
    repo = SQLiteCXRepository(":memory:")
    monkeypatch.setattr(test_whatsapp, "get_repository", lambda: repo)
    result = test_whatsapp.simulate_whatsapp_send(
        test_whatsapp.LocalWhatsAppSend(to="919999999999", text="test reply"),
        x_test_whatsapp_signature="local-secret",
    )
    assert result["provider"] == "local_mock"
    assert result["status"] == "sent"
    assert repo.list_audit_events(result["correlation_id"])[0]["event_type"] == "local_whatsapp_test_sent"


def test_local_whatsapp_manual_meta_send_uses_cloud_connector(monkeypatch):
    monkeypatch.setenv("WHATSAPP_LOCAL_TEST_MODE", "true")
    monkeypatch.setenv("WHATSAPP_TEST_SIGNATURE", "local-secret")
    monkeypatch.setenv("WHATSAPP_ACCESS_TOKEN", "meta-token")
    monkeypatch.setenv("WHATSAPP_PHONE_NUMBER_ID", "phone-id")
    repo = SQLiteCXRepository(":memory:")
    monkeypatch.setattr(test_whatsapp, "get_repository", lambda: repo)

    class FakeMetaConnector:
        def __init__(self, access_token, phone_number_id):
            assert access_token == "meta-token"
            assert phone_number_id == "phone-id"

        def send_text(self, to, text):
            return {"messages": [{"id": "wamid-meta-test"}], "to": to, "text": text}

    monkeypatch.setattr(test_whatsapp, "WhatsAppCloudConnector", FakeMetaConnector)
    result = test_whatsapp.simulate_whatsapp_send(
        test_whatsapp.LocalWhatsAppSend(to="919999999999", text="real Meta test", provider="meta"),
        x_test_whatsapp_signature="local-secret",
    )
    assert result["provider"] == "meta"
    assert result["provider_response"]["messages"][0]["id"] == "wamid-meta-test"


def test_whatsapp_status_webhook_updates_outbound_turn_and_admin_lookup(monkeypatch):
    from fastapi.testclient import TestClient

    from apps.api.main import app

    repo = SQLiteCXRepository(":memory:")

    class MetaSender:
        def send_text(self, to, text):
            return {"messages": [{"id": "wamid-meta-reply"}], "contacts": [{"wa_id": to}]}

    response = graph(repo, whatsapp=MetaSender()).run(whatsapp_message(metadata={"provider": "whatsapp_cloud"}))
    assert response.outbound_status == "sent"

    monkeypatch.setenv("WHATSAPP_APP_SECRET", "status-secret")
    monkeypatch.setenv("ADMIN_API_KEY", "status-admin-key")
    monkeypatch.setattr(integrations, "get_repository", lambda: repo)
    monkeypatch.setattr(whatsapp_admin, "get_repository", lambda: repo)

    payload = {
        "entry": [
            {
                "changes": [
                    {
                        "value": {
                            "statuses": [
                                {
                                    "id": "wamid-meta-reply",
                                    "status": "delivered",
                                    "timestamp": "1710000000",
                                    "recipient_id": "918555870077",
                                    "conversation": {"id": "conv-meta"},
                                    "pricing": {"category": "service"},
                                }
                            ]
                        }
                    }
                ]
            }
        ]
    }
    body = json.dumps(payload, separators=(",", ":")).encode()
    signature = "sha256=" + hmac.new(b"status-secret", body, hashlib.sha256).hexdigest()

    client = TestClient(app)
    webhook = client.post(
        "/integrations/whatsapp/webhook",
        content=body,
        headers={"x-hub-signature-256": signature, "content-type": "application/json"},
    )

    assert webhook.status_code == 202
    assert webhook.json()["statuses_received"] == 1
    assert webhook.json()["statuses"][0]["status"] == "delivered"

    admin = client.get(
        "/admin/whatsapp/delivery-statuses",
        params={"provider_message_id": "wamid-meta-reply"},
        headers={"x-admin-key": "status-admin-key"},
    )

    assert admin.status_code == 200
    assert admin.json()[0]["status"] == "delivered"

    outbound_turn = next(
        turn
        for turn in repo.get_conversation(response.conversation_id)["turns"]
        if turn["direction"] == "outbound"
    )
    assert outbound_turn["delivery_status"] == "delivered"
    assert outbound_turn["metadata"]["provider_message_id"] == "wamid-meta-reply"


class FakeCRM:
    def lookup_customer(self, channel, identifier):
        return CRMResult("synced", {"customer_id": "crm-customer-001", "segment": "premium"})

    def create_ticket(self, ticket, customer=None):
        return CRMResult(
            "synced",
            {
                "external_ticket_id": "BFSI-101",
                "external_ticket_url": "https://your-domain.atlassian.net/browse/BFSI-101",
            },
        )

    def add_comment(self, external_ticket_id, comment):
        assert external_ticket_id == "BFSI-101"
        return CRMResult("synced", {"comment_id": "comment-1"})

    def update_ticket_status(self, external_ticket_id, status):
        assert external_ticket_id == "BFSI-101"
        return CRMResult("synced", {"status": status})


def test_invalid_crm_url_is_recorded_without_raising(monkeypatch):
    from services.crm_service.client import CRMClient

    monkeypatch.setenv("CRM_PROVIDER", "jira")
    monkeypatch.setenv("CRM_BASE_URL", "Log in with Atlassian account")
    monkeypatch.setenv("CRM_API_TOKEN", "token")
    monkeypatch.setenv("CRM_USER_EMAIL", "support@example.com")
    monkeypatch.setenv("CRM_PROJECT_KEY", "SUP")

    result = CRMClient()._request("GET", "/rest/api/3/myself")

    assert result.status == "failed"
    assert "invalid URL" in result.error


def test_invalid_crm_url_does_not_block_whatsapp_reply(monkeypatch):
    monkeypatch.setenv("CRM_PROVIDER", "jira")
    monkeypatch.setenv("CRM_BASE_URL", "Log in with Atlassian account")
    monkeypatch.setenv("CRM_API_TOKEN", "token")
    monkeypatch.setenv("CRM_USER_EMAIL", "support@example.com")
    monkeypatch.setenv("CRM_PROJECT_KEY", "SUP")

    repo = SQLiteCXRepository(":memory:")
    sender = Recorder()
    response = graph(repo, whatsapp=sender).run(
        whatsapp_message(message_id="crm-failure-reply", text="unknown question xyz")
    )

    assert response.outbound_status == "sent"
    assert sender.sent
    ticket = repo.get_ticket(response.ticket_id)
    assert ticket["crm_sync_status"] == "failed"
    assert "invalid URL" in ticket["crm_sync_error"]


def test_ticket_jira_sync_and_lifecycle():
    repo = SQLiteCXRepository(":memory:")
    workflow = graph(repo, crm=FakeCRM())
    response = workflow.run(
        email_message(body="This is a terrible complaint. The service is unacceptable and I am extremely frustrated.")
    )
    ticket = repo.get_ticket(response.ticket_id)
    assert ticket["external_ticket_id"] == "BFSI-101"
    assert "atlassian.net" in ticket["external_ticket_url"]
    assert ticket["crm_sync_status"] == "synced"
    assert ticket["sla_due_at"]
    assert ticket["escalation_reason"].startswith("manual_review_required")

    manager = workflow.tickets
    comment = manager.add_comment(response.ticket_id, "Escalated to fraud team.")
    assert comment["details"]["crm_sync_status"] == "synced"
    updated = manager.update_status(response.ticket_id, TicketStatus.IN_PROGRESS)
    assert updated["status"] == "in_progress"
    assert [event["event_type"] for event in repo.list_ticket_events(response.ticket_id)] == [
        "ticket_created",
        "crm_sync_synced",
        "comment_added",
        "status_updated",
    ]


def test_ticket_priority_score_round_trips_and_never_leaks_to_customer_reply():
    """Phase 1: smart case prioritization. A negative-sentiment complaint should produce
    a non-zero priority score that survives persistence, and the internal scoring
    rationale must never appear in the customer-facing reply text (compliance guard)."""
    repo = SQLiteCXRepository(":memory:")
    workflow = graph(repo, crm=FakeCRM())
    response = workflow.run(
        email_message(body="This is a terrible complaint. The service is unacceptable and I am extremely frustrated.")
    )
    ticket = repo.get_ticket(response.ticket_id)
    assert ticket["priority_score"] > 0
    assert isinstance(ticket["priority_breakdown"], dict)
    assert ticket["priority_breakdown"]["total"] == ticket["priority_score"]

    assert "priority_score" not in response.message
    assert "priority_breakdown" not in response.message
    assert str(ticket["priority_score"]) not in response.message


# ── Admin / infrastructure tests ──────────────────────────────────────────────

def test_admin_ui_is_served():
    from fastapi.testclient import TestClient

    from apps.api.main import app

    response = TestClient(app).get("/admin-ui")
    assert response.status_code == 200
    assert "Ticket Operations" in response.text
    assert "email-simulate-form" in response.text


def test_email_admin_routes_are_protected(monkeypatch):
    from fastapi.testclient import TestClient

    from apps.api.main import app

    monkeypatch.setenv("ADMIN_API_KEY", "email-admin-key")
    monkeypatch.setenv("SMTP_HOST", "smtp.gmail.com")
    monkeypatch.setenv("SMTP_PORT", "587")
    monkeypatch.setenv("SMTP_USERNAME", "support@example.com")
    monkeypatch.setenv("SMTP_PASSWORD", "app-password")
    monkeypatch.setenv("SMTP_USE_TLS", "true")
    client = TestClient(app)

    assert client.get("/admin/email/status").status_code == 401
    response = client.get("/admin/email/status", headers={"x-admin-key": "email-admin-key"})

    assert response.status_code == 200
    assert response.json()["gmail_ready"] is True


def test_whatsapp_admin_status_reports_meta_readiness(monkeypatch):
    from fastapi.testclient import TestClient

    from apps.api.main import app

    monkeypatch.setenv("ADMIN_API_KEY", "whatsapp-admin-key")
    monkeypatch.setenv("WHATSAPP_ACCESS_TOKEN", "meta-token")
    monkeypatch.setenv("WHATSAPP_PHONE_NUMBER_ID", "phone-id")
    monkeypatch.setenv("WHATSAPP_BUSINESS_ACCOUNT_ID", "waba-id")
    monkeypatch.setenv("WHATSAPP_VERIFY_TOKEN", "verify-token")
    monkeypatch.setenv("WHATSAPP_APP_SECRET", "app-secret")
    client = TestClient(app)

    assert client.get("/admin/whatsapp/status").status_code == 401
    response = client.get("/admin/whatsapp/status", headers={"x-admin-key": "whatsapp-admin-key"})

    assert response.status_code == 200
    assert response.json()["meta_outbound_ready"] is True
    assert response.json()["meta_webhook_ready"] is True
    assert response.json()["business_account_id_configured"] is True


def test_orchestration_trace_records_explicit_agent_workflow():
    repo = SQLiteCXRepository(":memory:")
    # Non-account-specific loan question so this message passes customer validation and
    # proceeds through the full agent chain instead of being rejected as unregistered.
    response = graph(repo).run(whatsapp_message(text="What are the requirements for a personal loan?"))
    steps = [entry["step"] for entry in response.workflow_trace]
    # No open case for this customer, so check_has_open_case routes straight to the
    # Agent 1 chain — detect_ticket_action is skipped entirely rather than run and
    # discarded, since there is nothing to close.
    assert steps == [
        "receive_message",
        "resolve_identity",
        "load_conversation_context",
        "check_has_open_case",
        "classify_intent",
        "validate_customer",
        "retrieve_knowledge",
        "decide_resolution",
        "create_or_update_ticket",
        "send_outbound_reply",
        "persist_audit_events",
    ]
    events = repo.list_audit_events(response.correlation_id)
    agent_actions = [event for event in events if event["event_type"] == "workflow_step_completed"]
    assert len(agent_actions) == len(steps)
    assert events[-1]["event_type"] == "workflow_completed"


def test_orchestration_workflow_definition_is_admin_protected(monkeypatch):
    from fastapi.testclient import TestClient

    from apps.api.main import app

    monkeypatch.setenv("ADMIN_API_KEY", "workflow-test-admin-key")
    client = TestClient(app)
    assert client.get("/admin/orchestration/workflow").status_code == 401
    response = client.get(
        "/admin/orchestration/workflow",
        headers={"x-admin-key": "workflow-test-admin-key"},
    )
    assert response.status_code == 200
    assert response.json()["engine"] == "langgraph_state_graph"
    assert response.json()["framework"] == "LangGraph"
    assert len(response.json()["agents"]) == 4
    assert response.json()["agents"][0]["name"] == "intent_classification_agent"
    assert "groq" in response.json()["agents"][0]["execution"]


def test_hashing_embeddings_are_an_explicit_fallback(monkeypatch):
    monkeypatch.setenv("EMBEDDING_BACKEND", "hashing")
    embeddings = SemanticEmbeddings()
    assert embeddings.status()["active_backend"] == "hashing_fallback"
    assert len(embeddings.embed_query("What is my loan balance?")) == 384


# ── PII masking ────────────────────────────────────────────────────────────────

def test_pii_masker_detects_pan_aadhar_phone_email():
    from services.pii_service.masker import mask_text

    masked, mapping = mask_text(
        "My PAN is ABCDE1234F, Aadhar is 1234 5678 9012, "
        "call me at 9876543210 or email test.user@example.com"
    )
    assert "ABCDE1234F" not in masked
    assert "1234 5678 9012" not in masked
    assert "9876543210" not in masked
    assert "test.user@example.com" not in masked
    assert set(mapping.values()) == {
        "ABCDE1234F", "1234 5678 9012", "9876543210", "test.user@example.com",
    }


def test_pii_masker_detects_masked_aadhar_format():
    from services.pii_service.masker import mask_text

    masked, mapping = mask_text("Aadhar on file shows XXXX-XXXX-4321.")
    assert "XXXX-XXXX-4321" not in masked
    assert list(mapping.values()) == ["XXXX-XXXX-4321"]


def test_pii_masker_luhn_valid_card_masked_but_internal_ids_are_not():
    from services.pii_service.masker import mask_text

    masked, mapping = mask_text(
        "My card number is 4532015112830366. "
        "My loan LN001002 balance is 500000, account 40900000100008."
    )
    assert "4532015112830366" not in masked
    assert list(mapping.values()) == ["4532015112830366"]
    # Internal reference IDs and amounts are NOT PII and must stay untouched — the LLM
    # needs them to generate a useful answer.
    assert "LN001002" in masked
    assert "500000" in masked
    assert "40900000100008" in masked


def test_pii_masker_known_values_substitution_and_round_trip():
    from services.pii_service.masker import mask_text, unmask_text

    original = "Hi, this is Fathima Devasahayam, my phone is 7538870992."
    masked, mapping = mask_text(
        original, known_values={"name": "Fathima Devasahayam", "phone": "7538870992"}
    )
    assert "Fathima Devasahayam" not in masked
    assert "7538870992" not in masked
    assert unmask_text(masked, mapping) == original


def test_pii_masker_toggle_disables_masking(monkeypatch):
    from services.pii_service.masker import mask_text

    monkeypatch.setenv("PII_MASKING_ENABLED", "false")
    masked, mapping = mask_text("My PAN is ABCDE1234F")
    assert masked == "My PAN is ABCDE1234F"
    assert mapping == {}


def test_generate_answer_masks_pii_before_sending_to_groq_and_restores_name():
    from services.rag_service.groq_generator import GroqGenerator

    captured = {}

    class FakeGenerator(GroqGenerator):
        def _generate(self, system_prompt, user_prompt):
            captured["prompt"] = user_prompt
            # Simulate the LLM using the masked name placeholder in a personalized greeting.
            return {"text": "Hello [NAME_1], your loan LN001002 status is Active.",
                    "model": "test", "llm_used": True}

    ctx = {
        "channel": "whatsapp",
        "graph_context": {
            "customer_id": "CRN00010005",
            "name": "Fathima Devasahayam",
            "phone": "7538870992",
            "email": "fathimawork511@gmail.com",
            "city": "Amritsar",
            "loans": [{
                "loan_id": "LN001002", "loan_type": "Personal Loan", "status": "Active",
                "amount_inr": 17072.94, "next_step": "EMI overdue",
            }],
        },
    }
    result = FakeGenerator().generate_answer(
        "My PAN is ABCDE1234F, what is my loan status?", [], ctx
    )
    prompt = captured["prompt"]
    assert "ABCDE1234F" not in prompt
    assert "Fathima Devasahayam" not in prompt
    assert "[PAN_1]" in prompt
    assert "[NAME_1]" in prompt
    # Non-PII operational data the LLM needs must stay untouched.
    assert "LN001002" in prompt
    # The final answer shown to the customer has the real name restored.
    assert result["text"] == "Hello Fathima Devasahayam, your loan LN001002 status is Active."


def test_specific_scope_refines_open_other_ticket_instead_of_forking():
    """Omnichannel continuation: a vague dispute ("...:other" scope) followed by a
    specific follow-up ("...:card") on ANY channel must refine the open ticket,
    not create a duplicate (Sayantini email->web_chat split, 23 Jul 2026)."""
    from services.ticket_service.ticket_manager import TicketManager
    from shared.schemas.intents import Urgency
    from shared.schemas.messages import Channel, InboundMessage

    repo = SQLiteCXRepository(":memory:")
    manager = TicketManager(repo)

    def inbound(channel: Channel, text: str, msg_id: str) -> InboundMessage:
        return InboundMessage(
            channel=channel,
            channel_identifier="cust-identifier",
            text=text,
            provider="test",
            external_message_id=msg_id,
            correlation_id=msg_id,
        )

    # Real customer + conversation rows (tickets has FK constraints on both).
    first = inbound(Channel.EMAIL, "I want to dispute a transaction on my last statement.", "m1")
    customer = repo.resolve_customer(first)
    conversation = repo.get_or_create_conversation(customer["customer_id"])
    conv_id, cust_id = conversation["conversation_id"], customer["customer_id"]

    vague = manager.create_or_get_ticket(
        conv_id, cust_id,
        first,
        Intent.TRANSACTION_DISPUTE, Urgency.HIGH,
        escalation_reason="assisted_resolution_required:transaction_dispute",
    )
    assert vague.metadata["ticket_scope"] == "transaction_dispute:other"

    specific = manager.create_or_get_ticket(
        conv_id, cust_id,
        inbound(Channel.WEB_CHAT, 'The charge of Rs. 4,500 at "TechMart" on my Mastercard Classic card.', "m2"),
        Intent.TRANSACTION_DISPUTE, Urgency.HIGH,
        escalation_reason="assisted_resolution_required:transaction_dispute",
    )
    # Same ticket, scope upgraded, refinement audit-trailed.
    assert specific.ticket_id == vague.ticket_id
    assert specific.metadata["ticket_scope"] == "transaction_dispute:card"
    events = [e["event_type"] for e in repo.list_ticket_events(vague.ticket_id)]
    assert "ticket_scope_refined" in events

    # Guard rail: a DIFFERENT specific scope (UPI) is a distinct incident -> new ticket.
    upi = manager.create_or_get_ticket(
        conv_id, cust_id,
        inbound(Channel.WHATSAPP, "I also want to dispute a UPI payment of Rs. 900.", "m3"),
        Intent.TRANSACTION_DISPUTE, Urgency.HIGH,
        escalation_reason="assisted_resolution_required:transaction_dispute",
    )
    assert upi.ticket_id != vague.ticket_id
    assert upi.metadata["ticket_scope"] == "transaction_dispute:upi"

    # Idempotency: repeating the card details still lands on the refined ticket.
    repeat = manager.create_or_get_ticket(
        conv_id, cust_id,
        inbound(Channel.WEB_CHAT, "Again: the card charge at TechMart is the disputed one.", "m4"),
        Intent.TRANSACTION_DISPUTE, Urgency.HIGH,
        escalation_reason="assisted_resolution_required:transaction_dispute",
    )
    assert repeat.ticket_id == vague.ticket_id


# ── Tier-4 ticket referee (LLM matches a scope-unmatched message to an open ticket) ──

class _FakeRefereeGenerator:
    """Stands in for GroqGenerator: returns a scripted answer per call and
    records the prompts it was given."""

    def __init__(self, answers: list[str]):
        self.answers = list(answers)
        self.prompts: list[str] = []
        self.raise_error = False

    def _generate(self, system_prompt: str, user_prompt: str, operation: str = "llm_generation", metadata=None) -> dict:
        if self.raise_error:
            raise RuntimeError("llm down")
        self.prompts.append(user_prompt)
        answer = self.answers.pop(0) if self.answers else "NEW"
        return {"text": answer, "llm_used": True, "model": "fake"}


def _referee_fixture():
    """Repo + manager + the demo's first two turns: an email :other opener
    refined to :card by a web-chat follow-up (one open ticket)."""
    from services.ticket_service.ticket_manager import TicketManager
    from shared.schemas.intents import Urgency
    from shared.schemas.messages import Channel, InboundMessage

    repo = SQLiteCXRepository(":memory:")
    manager = TicketManager(repo)

    counter = {"n": 0}

    def send(channel: Channel, text: str):
        counter["n"] += 1
        msg = InboundMessage(
            channel=channel, channel_identifier="cust-identifier", text=text,
            provider="test", external_message_id=f"r{counter['n']}", correlation_id=f"r{counter['n']}",
        )
        return manager.create_or_get_ticket(
            conv_id, cust_id, msg, Intent.TRANSACTION_DISPUTE, Urgency.HIGH,
            escalation_reason="assisted_resolution_required:transaction_dispute",
        )

    first = InboundMessage(
        channel=Channel.EMAIL, channel_identifier="cust-identifier",
        text="Hi, I need help disputing a transaction on my statement. Can you please help.",
        provider="test", external_message_id="r0", correlation_id="r0",
    )
    customer = repo.resolve_customer(first)
    conversation = repo.get_or_create_conversation(customer["customer_id"])
    conv_id, cust_id = conversation["conversation_id"], customer["customer_id"]

    opener = manager.create_or_get_ticket(
        conv_id, cust_id, first, Intent.TRANSACTION_DISPUTE, Urgency.HIGH,
        escalation_reason="assisted_resolution_required:transaction_dispute",
    )
    refined = send(Channel.WEB_CHAT, "It's the Rs. 4,500 charge at TechMart on my Mastercard Classic card.")
    assert refined.ticket_id == opener.ticket_id
    assert refined.metadata["ticket_scope"] == "transaction_dispute:card"
    return repo, manager, send, opener


def test_referee_attaches_vague_cross_channel_followup_to_open_ticket():
    """The live 23 Jul demo failure: email ':other' follow-up after refinement
    to ':card' must land on the SAME ticket when the LLM referee confirms it."""
    from shared.schemas.messages import Channel

    repo, manager, send, opener = _referee_fixture()
    manager.generator = _FakeRefereeGenerator([opener.ticket_id])

    followup = send(Channel.EMAIL, "Hi, Any update on my transaction dispute request? This is urgent for me.")
    assert followup.ticket_id == opener.ticket_id
    events = [e["event_type"] for e in repo.list_ticket_events(opener.ticket_id)]
    assert "ticket_referee_attached" in events
    # The referee saw the candidate ticket in its prompt.
    assert opener.ticket_id in manager.generator.prompts[0]


def test_referee_new_verdict_forks():
    """LLM says NEW (a genuinely different matter, e.g. another merchant) ->
    new ticket, not an attach to the open ':card' one."""
    from shared.schemas.messages import Channel

    _, manager, send, opener = _referee_fixture()
    manager.generator = _FakeRefereeGenerator(["NEW"])
    gym = send(Channel.EMAIL, "I want to dispute a transaction - my gym charged me twice this month.")
    assert gym.ticket_id != opener.ticket_id


def test_referee_skipped_without_generator_forks():
    """No generator wired -> referee skipped -> vague follow-up forks (exact
    pre-referee behavior preserved)."""
    from shared.schemas.messages import Channel

    _, manager, send, opener = _referee_fixture()
    manager.generator = None
    no_llm = send(Channel.WHATSAPP, "Please give me the latest update on my complaint.")
    assert no_llm.ticket_id != opener.ticket_id


def test_referee_hallucinated_id_forks():
    """An answer outside the vetted candidate set is rejected -> new ticket."""
    from shared.schemas.messages import Channel

    _, manager, send, opener = _referee_fixture()
    manager.generator = _FakeRefereeGenerator(["tkt_definitely_not_real"])
    halluc = send(Channel.EMAIL, "Any news about my issue?")
    assert halluc.ticket_id != opener.ticket_id


def test_referee_llm_error_forks():
    """The LLM raising must fall back to a new ticket - doubt forks, never merges."""
    from shared.schemas.messages import Channel

    _, manager, send, opener = _referee_fixture()
    broken = _FakeRefereeGenerator([])
    broken.raise_error = True
    manager.generator = broken
    errored = send(Channel.WHATSAPP, "Following up on my pending matter please.")
    assert errored.ticket_id != opener.ticket_id


def test_referee_can_pick_older_ticket_among_multiple_candidates():
    """Two open dispute tickets (:card older, :upi newer): the referee's answer
    - not recency - decides, so naming the OLDER ticket attaches there."""
    from shared.schemas.messages import Channel

    repo, manager, send, opener = _referee_fixture()

    manager.generator = _FakeRefereeGenerator(["NEW"])
    upi = send(Channel.WHATSAPP, "I also want to dispute a UPI payment of Rs. 900 I never made.")
    assert upi.ticket_id != opener.ticket_id
    assert upi.metadata["ticket_scope"] == "transaction_dispute:upi"

    manager.generator = _FakeRefereeGenerator([opener.ticket_id])
    techmart = send(Channel.EMAIL, "What is happening with the TechMart one?")
    assert techmart.ticket_id == opener.ticket_id
    # Both candidates were offered to the referee.
    prompt = manager.generator.prompts[0]
    assert opener.ticket_id in prompt and upi.ticket_id in prompt
