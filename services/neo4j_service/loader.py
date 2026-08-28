"""Load bfsi.xlsx into Neo4j graph.

Sheets handled:
  Customer_Demographics → (:Customer) nodes + (:KYC) nodes
  Accounts              → (:Account) nodes + [:HAS_ACCOUNT] + [:PRODUCT_IS]
  Credit_Cards          → (:CreditCard) nodes + [:HAS_CREDIT_CARD] + [:PRODUCT_IS]
  Loans                 → (:Loan) nodes + [:HAS_LOAN] + [:PRODUCT_IS]
  Fixed_Deposits         → (:FixedDeposit) nodes + [:HAS_FD] + [:PRODUCT_IS]
  Policies               → (:Policy) nodes + [:HAS_POLICY]
  Claims                 → (:Claim) nodes + [:HAS_CLAIM] (from Customer and from Policy)
  Transactions           → (:Transaction) nodes + [:HAS_TRANSACTION]
  Charges_Penalties      → (:ChargePenalty) nodes + [:HAS_CHARGE]
  Interactions           → (:Interaction) nodes + [:HAS_INTERACTION] + [:HANDLED_BY],
                            plus (:ResolutionMemory) seed entries + [:CREATED_MEMORY]
                            (same property shape as the runtime writer)
  Products_Catalog       → (:Product) nodes
Runtime:
  (:Agent) nodes for AI and human handlers (always created)
"""

from pathlib import Path
import logging
import openpyxl

logger = logging.getLogger(__name__)

ROOT = Path(__file__).resolve().parents[2]
BFSI_XLSX = ROOT / "data" / "bfsi.xlsx"


def load_bfsi_data(client) -> dict:
    wb = openpyxl.load_workbook(BFSI_XLSX)
    counts: dict[str, int] = {}

    counts["customers"] = _load_customers(client, wb)
    counts["kyc"] = _load_kyc(client, wb)
    counts["products"] = _load_products(client, wb)
    counts["accounts"] = _load_accounts(client, wb)
    counts["credit_cards"] = _load_credit_cards(client, wb)
    counts["loans"] = _load_loans(client, wb)
    counts["fixed_deposits"] = _load_fixed_deposits(client, wb)
    counts["policies"] = _load_policies(client, wb)
    counts["claims"] = _load_claims(client, wb)
    counts["transactions"] = _load_transactions(client, wb)
    counts["charges"] = _load_charges(client, wb)
    counts["interactions"] = _load_interactions(client, wb)
    counts["agents"] = _load_agents(client)

    logger.info("bfsi_data_loaded", extra=counts)
    return counts


def _rows(wb, sheet: str) -> list[dict]:
    if sheet not in wb.sheetnames:
        return []
    ws = wb[sheet]
    headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(v is not None for v in row):
            rows.append(dict(zip(headers, row)))
    return rows


# ── Customers & KYC ──────────────────────────────────────────────────────────

def _load_customers(client, wb) -> int:
    rows = _rows(wb, "Customer_Demographics")
    for row in rows:
        client.write(
            """
            MERGE (c:Customer {customer_id: $customer_id})
            SET c.name = $name,
                c.age = $age,
                c.gender = $gender,
                c.occupation = $occupation,
                c.email = $email,
                c.secondary_email = $secondary_email,
                c.phone = $phone,
                c.alternate_mobile = $alternate_mobile,
                c.pan = $pan,
                c.aadhar = $aadhar,
                c.city = $city,
                c.state = $state,
                c.country = 'India',
                c.income_level = $income_level,
                c.segment = $segment,
                c.registration_date = $registration_date,
                c.vintage_months = $vintage_months,
                c.total_product_holding = $total_product_holding
            """,
            {
                "customer_id": str(row.get("CRN", "")),
                "name": str(row.get("Name") or ""),
                "age": row.get("Age") or 0,
                "gender": str(row.get("Gender") or ""),
                "occupation": str(row.get("Occupation") or ""),
                "email": str(row.get("Email1") or ""),
                "secondary_email": str(row.get("AlternateEmail") or ""),
                "phone": str(row.get("Mobile1") or ""),
                "alternate_mobile": str(row.get("AlternateMobile") or ""),
                "pan": str(row.get("PAN") or ""),
                "aadhar": str(row.get("Aadhar") or ""),
                "city": str(row.get("City") or ""),
                "state": str(row.get("State") or ""),
                "income_level": str(row.get("IncomeLevelApprox") or ""),
                "segment": str(row.get("Segment") or ""),
                "registration_date": str(row.get("AccountOpeningDate") or ""),
                "vintage_months": row.get("VintageMonths") or 0,
                "total_product_holding": str(row.get("TotalProductHolding") or ""),
            },
        )
    return len(rows)


def _load_kyc(client, wb) -> int:
    """Create one :KYC node per customer, status sourced from the sheet."""
    rows = _rows(wb, "Customer_Demographics")
    for row in rows:
        customer_id = str(row.get("CRN", ""))
        if not customer_id:
            continue
        client.write(
            """
            MERGE (k:KYC {customer_id: $customer_id})
            SET k.kyc_status   = $kyc_status,
                k.registered_at = $registered_at
            WITH k
            MATCH (c:Customer {customer_id: $customer_id})
            MERGE (c)-[:KYC_VERIFIED_BY]->(k)
            """,
            {
                "customer_id": customer_id,
                "kyc_status": str(row.get("KYCStatus") or "Pending"),
                "registered_at": str(row.get("AccountOpeningDate") or ""),
            },
        )
    return len(rows)


# ── Accounts ─────────────────────────────────────────────────────────────────

def _load_accounts(client, wb) -> int:
    rows = _rows(wb, "Accounts")
    for row in rows:
        account_number = str(row.get("AccountNumber", ""))
        customer_id = str(row.get("CRN", ""))
        account_type = str(row.get("AccountType") or "")
        client.write(
            """
            MERGE (a:Account {account_number: $account_number})
            SET a.account_category = $account_category,
                a.account_type = $account_type,
                a.account_sub_type = $account_sub_type,
                a.opening_date = $opening_date,
                a.status = $status,
                a.branch = $branch,
                a.ifsc = $ifsc,
                a.avg_monthly_balance = $avg_monthly_balance,
                a.min_balance_required = $min_balance_required,
                a.currency = $currency
            WITH a
            MATCH (c:Customer {customer_id: $customer_id})
            MERGE (c)-[:HAS_ACCOUNT]->(a)
            WITH a
            MATCH (p:Product)
            WHERE p.category = $account_type
              AND p.product_type IN ['SavingsAccount', 'CurrentAccount']
            MERGE (a)-[:PRODUCT_IS]->(p)
            """,
            {
                "account_number": account_number,
                "customer_id": customer_id,
                "account_category": str(row.get("AccountCategory") or ""),
                "account_type": account_type,
                "account_sub_type": str(row.get("AccountSubType") or ""),
                "opening_date": str(row.get("OpeningDate") or ""),
                "status": str(row.get("Status") or ""),
                "branch": str(row.get("Branch") or ""),
                "ifsc": str(row.get("IFSC") or ""),
                "avg_monthly_balance": row.get("AvgMonthlyBalance") or 0,
                "min_balance_required": row.get("MinBalanceRequired") or 0,
                "currency": str(row.get("Currency") or "INR"),
            },
        )
    return len(rows)


# ── Credit Cards ─────────────────────────────────────────────────────────────

def _load_credit_cards(client, wb) -> int:
    rows = _rows(wb, "Credit_Cards")
    for row in rows:
        card_id = str(row.get("CardID", ""))
        customer_id = str(row.get("CRN", ""))
        card_variant = str(row.get("CardVariant") or "")
        client.write(
            """
            MERGE (cc:CreditCard {card_id: $card_id})
            SET cc.account_number = $account_number,
                cc.credit_limit = $credit_limit,
                cc.card_network = $card_network,
                cc.card_variant = $card_variant,
                cc.balance_due = $balance_due,
                cc.min_amount_due = $min_amount_due,
                cc.total_amount_due = $total_amount_due,
                cc.statement_date = $statement_date,
                cc.payment_due_date = $payment_due_date,
                cc.dpd = $dpd,
                cc.interest_rate = $interest_rate,
                cc.penalty_details = $penalty_details,
                cc.reward_points_balance = $reward_points_balance,
                cc.reward_points_expiry = $reward_points_expiry,
                cc.chargeback_flag = $chargeback_flag,
                cc.chargeback_reason = $chargeback_reason,
                cc.fraud_flag = $fraud_flag,
                cc.fraud_type = $fraud_type
            WITH cc
            MATCH (c:Customer {customer_id: $customer_id})
            MERGE (c)-[:HAS_CREDIT_CARD]->(cc)
            WITH cc
            MATCH (p:Product)
            WHERE p.product_type = 'CreditCard' AND p.category = $card_variant
            MERGE (cc)-[:PRODUCT_IS]->(p)
            """,
            {
                "card_id": card_id,
                "customer_id": customer_id,
                "account_number": str(row.get("AccountNumber") or ""),
                "credit_limit": row.get("CreditLimit") or 0,
                "card_network": str(row.get("CardNetwork") or ""),
                "card_variant": card_variant,
                "balance_due": row.get("BalanceDue") or 0,
                "min_amount_due": row.get("MinAmountDue") or 0,
                "total_amount_due": row.get("TotalAmountDue") or 0,
                "statement_date": str(row.get("StatementDate") or ""),
                "payment_due_date": str(row.get("PaymentDueDate") or ""),
                "dpd": row.get("DPD") or 0,
                "interest_rate": row.get("InterestRate") or 0,
                "penalty_details": str(row.get("PenaltyDetails") or "None"),
                "reward_points_balance": row.get("RewardPointsBalance") or 0,
                "reward_points_expiry": str(row.get("RewardPointsExpiry") or ""),
                "chargeback_flag": bool(row.get("ChargebackFlag")),
                "chargeback_reason": str(row.get("ChargebackReason") or ""),
                "fraud_flag": bool(row.get("FraudFlag")),
                "fraud_type": str(row.get("FraudType") or ""),
            },
        )
    return len(rows)


# ── Loans ────────────────────────────────────────────────────────────────────

def _load_loans(client, wb) -> int:
    rows = _rows(wb, "Loans")
    for row in rows:
        customer_id = str(row.get("CRN", ""))
        loan_id = str(row.get("LoanID", ""))
        loan_type = str(row.get("LoanType") or "")
        client.write(
            """
            MERGE (l:Loan {loan_id: $loan_id})
            SET l.account_number = $account_number,
                l.loan_type = $loan_type,
                l.collateral_type = $collateral_type,
                l.tenure_months = $tenure_months,
                l.principal_amount = $principal_amount,
                l.interest_rate = $interest_rate,
                l.status = $status,
                l.last_updated = $last_updated,
                l.amount_inr = $amount_inr,
                l.next_step = $next_step,
                l.dpd = $dpd,
                l.times_90_plus_dpd = $times_90_plus_dpd,
                l.times_60_plus_dpd = $times_60_plus_dpd,
                l.penalty_details = $penalty_details,
                l.emis_paid = $emis_paid,
                l.emis_pending = $emis_pending,
                l.total_emis = $total_emis
            WITH l
            MATCH (c:Customer {customer_id: $customer_id})
            MERGE (c)-[:HAS_LOAN]->(l)
            WITH l
            MATCH (p:Product)
            WHERE p.product_type = 'Loan'
              AND (p.category = $loan_type OR p.name CONTAINS $loan_type)
            MERGE (l)-[:PRODUCT_IS]->(p)
            """,
            {
                "loan_id": loan_id,
                "customer_id": customer_id,
                "account_number": str(row.get("AccountNumber") or ""),
                "loan_type": loan_type,
                "collateral_type": str(row.get("CollateralType") or "None"),
                "tenure_months": row.get("TenureMonths") or 0,
                "principal_amount": row.get("PrincipalAmount") or 0,
                "interest_rate": row.get("InterestRate") or 0,
                "status": str(row.get("Status") or ""),
                "last_updated": str(row.get("LastUpdatedDate") or ""),
                "amount_inr": row.get("BalanceDue") or 0,
                "next_step": str(row.get("NextStep") or ""),
                "dpd": row.get("DPD") or 0,
                "times_90_plus_dpd": row.get("Times90PlusDPD") or 0,
                "times_60_plus_dpd": row.get("Times60PlusDPD") or 0,
                "penalty_details": str(row.get("PenaltyDetails") or "None"),
                "emis_paid": row.get("EMIsPaid") or 0,
                "emis_pending": row.get("EMIsPending") or 0,
                "total_emis": row.get("TotalEMIs") or 0,
            },
        )
    return len(rows)


# ── Fixed Deposits ───────────────────────────────────────────────────────────

def _load_fixed_deposits(client, wb) -> int:
    rows = _rows(wb, "Fixed_Deposits")
    for row in rows:
        customer_id = str(row.get("CRN", ""))
        fd_id = str(row.get("FDID", ""))
        client.write(
            """
            MERGE (fd:FixedDeposit {fd_id: $fd_id})
            SET fd.account_number = $account_number,
                fd.principal_amount = $principal_amount,
                fd.interest_rate = $interest_rate,
                fd.tenure_months = $tenure_months,
                fd.booking_date = $booking_date,
                fd.maturity_date = $maturity_date,
                fd.maturity_amount = $maturity_amount,
                fd.auto_renewal = $auto_renewal,
                fd.status = $status
            WITH fd
            MATCH (c:Customer {customer_id: $customer_id})
            MERGE (c)-[:HAS_FD]->(fd)
            WITH fd
            MATCH (p:Product)
            WHERE p.product_type = 'FD' AND p.category = 'Regular'
            MERGE (fd)-[:PRODUCT_IS]->(p)
            """,
            {
                "fd_id": fd_id,
                "customer_id": customer_id,
                "account_number": str(row.get("AccountNumber") or ""),
                "principal_amount": row.get("PrincipalAmount") or 0,
                "interest_rate": row.get("InterestRate") or 0,
                "tenure_months": row.get("TenureMonths") or 0,
                "booking_date": str(row.get("BookingDate") or ""),
                "maturity_date": str(row.get("MaturityDate") or ""),
                "maturity_amount": row.get("MaturityAmount") or 0,
                "auto_renewal": str(row.get("AutoRenewal") or "N"),
                "status": str(row.get("Status") or ""),
            },
        )
    return len(rows)


# ── Policies & Claims ────────────────────────────────────────────────────────

def _load_policies(client, wb) -> int:
    rows = _rows(wb, "Policies")
    for row in rows:
        customer_id = str(row.get("CRN", ""))
        policy_id = str(row.get("PolicyID", ""))
        client.write(
            """
            MERGE (p:Policy {policy_id: $policy_id})
            SET p.policy_type = $policy_type,
                p.customer_id = $customer_id,
                p.premium_inr = $premium_inr,
                p.coverage_inr = $coverage_inr,
                p.premium_frequency = $premium_frequency,
                p.start_date = $start_date,
                p.maturity_date = $maturity_date,
                p.next_premium_due = $next_premium_due,
                p.status = $status,
                p.nominee_name = $nominee_name
            WITH p
            MATCH (c:Customer {customer_id: $customer_id})
            MERGE (c)-[:HAS_POLICY]->(p)
            """,
            {
                "policy_id": policy_id,
                "customer_id": customer_id,
                "policy_type": str(row.get("PolicyType") or ""),
                "premium_inr": row.get("PremiumAmount") or 0,
                "coverage_inr": row.get("CoverageAmount") or 0,
                "premium_frequency": str(row.get("PremiumFrequency") or ""),
                "start_date": str(row.get("StartDate") or ""),
                "maturity_date": str(row.get("MaturityDate") or ""),
                "next_premium_due": str(row.get("NextPremiumDueDate") or ""),
                "status": str(row.get("Status") or ""),
                "nominee_name": str(row.get("NomineeName") or ""),
            },
        )
    return len(rows)


def _load_claims(client, wb) -> int:
    rows = _rows(wb, "Claims")
    for row in rows:
        customer_id = str(row.get("CRN", ""))
        claim_id = str(row.get("ClaimID", ""))
        policy_id = str(row.get("PolicyID", ""))
        client.write(
            """
            MERGE (cl:Claim {claim_id: $claim_id})
            SET cl.policy_id = $policy_id,
                cl.claim_type = $claim_type,
                cl.status = $status,
                cl.last_updated = $last_updated,
                cl.amount_claimed_inr = $amount_claimed,
                cl.amount_approved_inr = $amount_approved,
                cl.reason = $reason
            WITH cl
            MATCH (c:Customer {customer_id: $customer_id})
            MERGE (c)-[:HAS_CLAIM]->(cl)
            WITH cl
            MATCH (p:Policy {policy_id: $policy_id})
            MERGE (p)-[:HAS_CLAIM]->(cl)
            """,
            {
                "claim_id": claim_id,
                "customer_id": customer_id,
                "policy_id": policy_id,
                "claim_type": str(row.get("ClaimType") or ""),
                "status": str(row.get("ClaimStatus") or ""),
                "last_updated": str(row.get("LastUpdatedDate") or ""),
                "amount_claimed": row.get("AmountClaimed") or 0,
                "amount_approved": row.get("AmountApproved") or 0,
                "reason": str(row.get("ReasonForStatus") or ""),
            },
        )
    return len(rows)


# ── Transactions & Charges ───────────────────────────────────────────────────

def _load_transactions(client, wb) -> int:
    rows = _rows(wb, "Transactions")
    for row in rows:
        customer_id = str(row.get("CRN", ""))
        txn_id = str(row.get("TxnID", ""))
        client.write(
            """
            MERGE (t:Transaction {txn_id: $txn_id})
            SET t.account_number = $account_number,
                t.txn_date = $txn_date,
                t.amount = $amount,
                t.txn_type = $txn_type,
                t.channel = $channel,
                t.beneficiary_account = $beneficiary_account,
                t.beneficiary_name = $beneficiary_name,
                t.status = $status,
                t.failure_reason = $failure_reason,
                t.narration = $narration
            WITH t
            MATCH (c:Customer {customer_id: $customer_id})
            MERGE (c)-[:HAS_TRANSACTION]->(t)
            """,
            {
                "txn_id": txn_id,
                "customer_id": customer_id,
                "account_number": str(row.get("AccountNumber") or ""),
                "txn_date": str(row.get("TxnDate") or ""),
                "amount": row.get("Amount") or 0,
                "txn_type": str(row.get("TxnType") or ""),
                "channel": str(row.get("Channel") or ""),
                "beneficiary_account": str(row.get("BeneficiaryAccount") or ""),
                "beneficiary_name": str(row.get("BeneficiaryName") or ""),
                "status": str(row.get("Status") or ""),
                "failure_reason": str(row.get("FailureReason") or ""),
                "narration": str(row.get("Narration") or ""),
            },
        )
    return len(rows)


def _load_charges(client, wb) -> int:
    rows = _rows(wb, "Charges_Penalties")
    for row in rows:
        customer_id = str(row.get("CRN", ""))
        charge_id = str(row.get("ChargeID", ""))
        client.write(
            """
            MERGE (ch:ChargePenalty {charge_id: $charge_id})
            SET ch.account_number = $account_number,
                ch.charge_type = $charge_type,
                ch.amount = $amount,
                ch.charge_date = $charge_date,
                ch.reason = $reason,
                ch.reversal_status = $reversal_status
            WITH ch
            MATCH (c:Customer {customer_id: $customer_id})
            MERGE (c)-[:HAS_CHARGE]->(ch)
            """,
            {
                "charge_id": charge_id,
                "customer_id": customer_id,
                "account_number": str(row.get("AccountNumber") or ""),
                "charge_type": str(row.get("ChargeType") or ""),
                "amount": row.get("Amount") or 0,
                "charge_date": str(row.get("ChargeDate") or ""),
                "reason": str(row.get("Reason") or ""),
                "reversal_status": str(row.get("ReversalStatus") or ""),
            },
        )
    return len(rows)


# ── Interactions & Resolution Memory ─────────────────────────────────────────

def _load_interactions(client, wb) -> int:
    """Seed historical conversation/resolution data.

    Writes the exact same (:Interaction) property shape as
    services/neo4j_service/writer.py::update_interaction_resolution(), so bulk-loaded
    history is indistinguishable from history written by the live app. Rows also
    seed (:ResolutionMemory) entries with the row's Verified flag, matching
    search_resolution_memory()'s verified-only read filter.
    """
    rows = _rows(wb, "Interactions")
    for row in rows:
        customer_id = str(row.get("CRN", ""))
        conversation_id = str(row.get("ConversationID", ""))
        intent = str(row.get("Intent") or "")
        product_ref = str(row.get("ProductRef") or "general")
        handled_by = str(row.get("HandledBy") or "AI_GROQ")
        verified = str(row.get("Verified") or "N").upper() == "Y"

        client.write(
            """
            MERGE (i:Interaction {conversation_id: $conversation_id})
            SET i.channel = $channel,
                i.message = $message,
                i.resolution = $resolution,
                i.intent = $intent,
                i.sentiment = $sentiment,
                i.urgency = $urgency,
                i.product_ref = $product_ref,
                i.status = 'closed',
                i.handled_by = $handled_by,
                i.created_at = $created_at,
                i.updated_at = $updated_at
            WITH i
            MATCH (c:Customer {customer_id: $customer_id})
            MERGE (c)-[:HAS_INTERACTION]->(i)
            WITH i
            MATCH (a:Agent {agent_id: $handled_by})
            MERGE (i)-[:HANDLED_BY]->(a)
            """,
            {
                "conversation_id": conversation_id,
                "customer_id": customer_id,
                "channel": str(row.get("Channel") or ""),
                "message": str(row.get("MessageText") or ""),
                "resolution": str(row.get("ResolutionText") or ""),
                "intent": intent,
                "sentiment": str(row.get("Sentiment") or "neutral"),
                "urgency": str(row.get("Urgency") or "low"),
                "product_ref": product_ref,
                "handled_by": handled_by,
                "created_at": str(row.get("CreatedAt") or ""),
                "updated_at": str(row.get("UpdatedAt") or ""),
            },
        )

        client.write(
            """
            MATCH (i:Interaction {conversation_id: $conversation_id})
            MERGE (rm:ResolutionMemory {memory_key: $memory_key})
            ON CREATE SET
                rm.id = $mem_id,
                rm.intent_type = $intent_type,
                rm.product_id = $product_id,
                rm.query_pattern = i.message,
                rm.resolution_text = i.resolution,
                rm.verified = $verified,
                rm.times_reused = 0,
                rm.created_at = i.created_at
            ON MATCH SET
                rm.intent_type = $intent_type,
                rm.resolution_text = i.resolution,
                rm.verified = $verified,
                rm.updated_at = i.updated_at
            MERGE (i)-[:CREATED_MEMORY]->(rm)
            """,
            {
                "conversation_id": conversation_id,
                "product_id": product_ref,
                "intent_type": intent,
                # Must match the runtime key in writer.update_interaction_resolution, or a
                # seeded memory can never be found by search_resolution_memory. Seed rows
                # carry no ticket_scope, so they take the "<intent>:general" fallback.
                "memory_key": f"{intent or 'unknown'}:general",
                "mem_id": "RESMEM-" + conversation_id,
                "verified": verified,
            },
        )
    return len(rows)


# ── Products ─────────────────────────────────────────────────────────────────

def _load_products(client, wb) -> int:
    rows = _rows(wb, "Products_Catalog")
    for row in rows:
        client.write(
            """
            MERGE (p:Product {product_id: $product_id})
            SET p.name = $name,
                p.product_type = $product_type,
                p.category = $category,
                p.description = $description,
                p.key_features = $key_features,
                p.eligibility = $eligibility
            """,
            {
                "product_id": str(row.get("ProductID") or ""),
                "name": str(row.get("ProductName") or ""),
                "product_type": str(row.get("ProductType") or ""),
                "category": str(row.get("Category") or ""),
                "description": str(row.get("Description") or ""),
                "key_features": str(row.get("KeyFeatures") or ""),
                "eligibility": str(row.get("EligibilityCriteria") or ""),
            },
        )
    return len(rows)


# ── Agents ───────────────────────────────────────────────────────────────────

def _load_agents(client) -> int:
    """Create default Agent nodes for AI handler and stub human handler."""
    agents = [
        {
            "agent_id": "AI_GROQ",
            "agent_type": "ai",
            "name": "InboxIQ AI",
            "model": "llama-3.1-8b-instant",
        },
        {
            "agent_id": "HUMAN_SR",
            "agent_type": "human",
            "name": "Support Representative",
            "model": "",
        },
    ]
    for agent in agents:
        client.write(
            """
            MERGE (a:Agent {agent_id: $agent_id})
            SET a.agent_type = $agent_type,
                a.name       = $name,
                a.model      = $model
            """,
            agent,
        )
    return len(agents)


if __name__ == "__main__":
    import sys
    sys.path.insert(0, str(ROOT))
    from services.neo4j_service.client import Neo4jClient
    client = Neo4jClient()
    result = load_bfsi_data(client)
    print("Loaded:", result)
    client.close()
